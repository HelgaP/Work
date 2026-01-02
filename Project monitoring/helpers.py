# вспомогательные функции
# форматирование ширины колонок

from openpyxl import load_workbook 
from openpyxl.drawing.image import Image

def formatting(path_to_file, option="general", path_to_dash=''): # if business = True, use short verison
    
    wb = load_workbook(path_to_file)
    ws3 = 0
    ws4 = 0

    if option == "general" or option == "business":
        
        ws1 = wb['Данные']
        ws2 = wb['Глоссарий']

        # лист "Глоссарий"
        ws2.column_dimensions['A'].width = 26.5
        ws2.column_dimensions['B'].width = 65.5
        ws2.column_dimensions['C'].width = 36 

        if option == "general":

                       #index = [ 1,  2,  3,  4,  5,  6,  7,  8,  9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21,22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42]
            length_of_columns = [25, 23, 12, 24, 24, 11, 16, 32, 11, 29, 31, 17, 19, 19, 14, 19, 25, 18, 10, 26, 20, 13, 10, 16, 16, 14, 26, 23, 14, 20, 47, 17, 30, 24, 14, 15, 20, 17, 19, 19, 25, 25, 8, 11]
            columns_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR']
        
        else:

            if path_to_dash != '':
                ws = wb.create_sheet("Дашборд")
                img = Image(path_to_dash)
                ws.add_image(img, 'A1')

                ws3 = wb['Строки со смещением']
                ws4 = wb['Файл для клиента']

                           #index = [ 1,  2,  3,  4,  5,  6,  7, 8,  9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19
                length_of_columns = [18, 20, 23, 22, 11, 11, 8, 30, 11, 14, 47, 26, 18, 26, 23, 26, 23, 10, 7]
                columns_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']

                #index =                   [ 1,  2,  3,  4,  5,  6,  7, 8,  9, 10, 11, 12, 13, 14, 15, 16
                length_of_columns_client = [18, 20, 23, 11, 11, 8, 30, 11, 14, 47, 26, 18, 26, 23, 23, 10]
                columns_letters_client = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

                for i in range(len(columns_letters)):
                    ws3.column_dimensions[columns_letters[i]].width = length_of_columns[i]

                for i in range(len(columns_letters_client)):
                    ws4.column_dimensions[columns_letters_client[i]].width = length_of_columns_client[i]

    elif option == 'cpk':

        ws1 = wb.active
                   #index = [ 1,  2,  3,  4,  5,  6,  7
        length_of_columns = [20, 15, 33, 22, 11, 8, 22]
        columns_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

    elif option == 'purch_lines':

        ws1 = wb.active
                   #index = [ 1,  2,  3
        length_of_columns = [30, 26, 29]
        columns_letters = ['A', 'B', 'C']
  
    
    # лист "Данные"
    for i in range(len(columns_letters)):
        ws1.column_dimensions[columns_letters[i]].width = length_of_columns[i]
    
    wb.save(path_to_file)
